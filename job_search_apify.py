import os
import re
import time
import smtplib
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from urllib.parse import quote_plus
from apify_client import ApifyClient

# -----------------------------------------------
# CONFIG
# -----------------------------------------------
KEYWORDS = ["Data Scientist", "Machine Learning Engineer", "AI Engineer"]
LOCATION = "India"

APIFY_TOKEN        = os.environ["APIFY_TOKEN"]
GMAIL_USER         = os.environ["GMAIL_USER"]
GMAIL_APP_PASSWORD = os.environ["GMAIL_APP_PASSWORD"]

_recipient = os.environ.get("RECIPIENT_EMAIL", "").strip()
RECIPIENT_EMAIL = _recipient if _recipient else GMAIL_USER

# columns to keep in the final Excel (in this order)
KEEP_COLUMNS = ["Platform", "Keyword", "Title", "Company", "Location", "Posted At", "Apply Link"]

# -----------------------------------------------
# ACTOR IDs
# -----------------------------------------------
ACTORS = {
    "Naukri":   "memo23/naukri-scraper",
    "LinkedIn": "valig/linkedin-jobs-scraper",
    "Indeed":   "valig/indeed-jobs-scraper",
    "Foundit":  "easyapi/foundit-jobs-scraper",
    "Shine":    "easyapi/shine-com-jobs-scraper",
}

# Per-platform result limits
LIMITS = {
    "Naukri":   30,
    "LinkedIn": 30,
    "Indeed":   20,
    "Foundit":  20,
    "Shine":    20,
}

# -----------------------------------------------
# BUILD LINKEDIN SEARCH URL
# f_TPR=r604800 = posted in last 7 days (604800 seconds)
# -----------------------------------------------
def build_linkedin_url(keyword, location):
    kw  = quote_plus(keyword)
    loc = quote_plus(location)
    return f"https://www.linkedin.com/jobs/search/?keywords={kw}&location={loc}&f_TPR=r604800&position=1&pageNum=0"

# -----------------------------------------------
# ACTOR INPUTS
# -----------------------------------------------
def get_actor_input(platform, keyword, location):
    limit = LIMITS.get(platform, 20)

    if platform == "Naukri":
        return {
            "keyword":  keyword,
            "location": location,
            "limit":    limit,
        }
    elif platform == "LinkedIn":
        return {
            "urls":  [build_linkedin_url(keyword, location)],
            "count": limit,
        }
    elif platform == "Indeed":
        return {
            "title":      keyword,
            "location":   location,
            "country":    "in",
            "limit":      limit,
            "datePosted": "7",
        }
    elif platform == "Foundit":
        return {
            "searchKeywords": keyword,
            "location":       location,
            "maxResults":     limit,
        }
    elif platform == "Shine":
        return {
            "searchKeywords": keyword,
            "location":       location,
            "maxResults":     limit,
        }

# -----------------------------------------------
# NORMALIZE
# LinkedIn job URL is built from the job ID because
# the actor does not return a direct URL field.
# -----------------------------------------------
def get_linkedin_url(raw):
    for field in ["jobUrl", "url", "applyUrl", "applyLink", "link", "jobLink", "externalApplyLink"]:
        val = raw.get(field)
        if val and isinstance(val, str) and val.startswith("http"):
            return val
    # fallback: construct from job ID
    job_id = raw.get("id") or raw.get("jobId") or raw.get("entityUrn")
    if job_id:
        if isinstance(job_id, str) and ":" in job_id:
            job_id = job_id.split(":")[-1]
        return f"https://www.linkedin.com/jobs/view/{job_id}"
    return None

def normalize_job(raw, platform, keyword):
    job = {
        "Platform":   platform,
        "Keyword":    keyword,
        "Title":      None,
        "Company":    None,
        "Location":   None,
        "Posted At":  None,
        "Apply Link": None,
    }

    if platform == "Naukri":
        job["Title"]      = raw.get("title") or raw.get("jobTitle")
        job["Company"]    = raw.get("companyName") or raw.get("company")
        job["Location"]   = raw.get("location") or raw.get("city")
        job["Posted At"]  = raw.get("postedDate") or raw.get("datePosted") or raw.get("createdAt")
        job["Apply Link"] = raw.get("jobUrl") or raw.get("url") or raw.get("applyLink")

    elif platform == "LinkedIn":
        job["Title"]      = raw.get("title") or raw.get("jobTitle")
        job["Company"]    = raw.get("company") or raw.get("companyName")
        job["Location"]   = raw.get("location") or raw.get("jobLocation")
        job["Posted At"]  = raw.get("postedAt") or raw.get("datePosted")
        job["Apply Link"] = get_linkedin_url(raw)

    elif platform == "Indeed":
        job["Title"]      = raw.get("positionName") or raw.get("title")
        job["Company"]    = raw.get("company")
        job["Location"]   = raw.get("location")
        job["Posted At"]  = raw.get("postedAt") or raw.get("date")
        job["Apply Link"] = raw.get("url") or raw.get("jobUrl")

    elif platform == "Foundit":
        job["Title"]      = raw.get("jobTitle") or raw.get("title")
        job["Company"]    = raw.get("companyName") or raw.get("company")
        job["Location"]   = raw.get("location") or raw.get("city")
        job["Posted At"]  = raw.get("postedAt") or raw.get("datePosted") or raw.get("createdAt")
        job["Apply Link"] = raw.get("jobUrl") or raw.get("url")

    elif platform == "Shine":
        job["Title"]      = raw.get("jobTitle") or raw.get("title")
        job["Company"]    = raw.get("companyName") or raw.get("company")
        job["Location"]   = raw.get("location") or raw.get("city")
        job["Posted At"]  = raw.get("postedAt") or raw.get("datePosted") or raw.get("createdAt")
        job["Apply Link"] = raw.get("jobUrl") or raw.get("url")

    return job

# -----------------------------------------------
# FILTER - last 7 days
# Also handles relative strings like "3 days ago", "1 week ago"
# -----------------------------------------------
def is_within_7d(posted_at_str):
    if not posted_at_str:
        return True
    s = str(posted_at_str).strip().lower()
    cutoff = datetime.now(timezone.utc) - timedelta(days=7)

    # Handle relative strings returned by some scrapers
    m = re.match(r"(\d+)\s*(minute|hour|day|week|month)s?\s*ago", s)
    if m:
        n, unit = int(m.group(1)), m.group(2)
        delta = {"minute": timedelta(minutes=n), "hour": timedelta(hours=n),
                 "day": timedelta(days=n), "week": timedelta(weeks=n),
                 "month": timedelta(days=n * 30)}[unit]
        return (datetime.now(timezone.utc) - delta) >= cutoff

    formats = [
        "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt >= cutoff
        except ValueError:
            continue
    return True

# -----------------------------------------------
# FETCH JOBS
# -----------------------------------------------
def fetch_all_jobs():
    client   = ApifyClient(APIFY_TOKEN)
    all_jobs = []

    for keyword in KEYWORDS:
        for platform, actor_id in ACTORS.items():
            print(f"Running {platform} for keyword: {keyword}")
            try:
                run        = client.actor(actor_id).call(run_input=get_actor_input(platform, keyword, LOCATION))
                time.sleep(2)
                dataset_id = (
                    getattr(run, "default_dataset_id", None)
                    or (run.get("defaultDatasetId") if isinstance(run, dict) else None)
                )
                if not dataset_id:
                    print(f"  No dataset returned, skipping.")
                    continue
                items = list(client.dataset(dataset_id).iterate_items())
                print(f"  Got {len(items)} raw results")
                for raw in items:
                    job = normalize_job(raw, platform, keyword)
                    if is_within_7d(job["Posted At"]):
                        all_jobs.append(job)
            except Exception as e:
                print(f"  Error: {e}")
                continue

    if not all_jobs:
        return pd.DataFrame()

    df = pd.DataFrame(all_jobs)

    for col in df.columns:
        df[col] = df[col].apply(lambda x: str(x) if isinstance(x, (dict, list)) else x)

    df = df.drop_duplicates(subset=["Title", "Company", "Location"])
    df = df.reset_index(drop=True)

    df = df.groupby("Platform", group_keys=False).apply(
        lambda x: x.head(LIMITS.get(x.name, 20))
    )
    df = df.reset_index(drop=True)

    return df

# -----------------------------------------------
# BUILD EXCEL
# - only KEEP_COLUMNS
# - Apply Link shows "Apply" as a clickable hyperlink
# - minimal fixed column widths
# - header row frozen
# -----------------------------------------------
def build_excel(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    df = df[[c for c in KEEP_COLUMNS if c in df.columns]]

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    link_font   = Font(name="Arial", color="0563C1", underline="single", size=10)
    body_font   = Font(name="Arial", size=10)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    link_col_idx = list(df.columns).index("Apply Link") + 1 if "Apply Link" in df.columns else None

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx == link_col_idx and value and str(value).startswith("http"):
                cell.value     = "Apply"
                cell.hyperlink = str(value)
                cell.font      = link_font
            else:
                cell.value = value if value and str(value) not in ("None", "nan") else ""
                cell.font  = body_font
            cell.alignment = Alignment(vertical="center")

    col_widths = {
        "Platform":   12,
        "Keyword":    22,
        "Title":      40,
        "Company":    22,
        "Location":   18,
        "Posted At":  18,
        "Apply Link": 10,
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# -----------------------------------------------
# SEND EMAIL
# -----------------------------------------------
def send_email(df, excel_bytes):
    today            = datetime.now().strftime("%Y-%m-%d")
    total            = len(df)
    platform_summary = "\n".join([f"  - {k}: {v} jobs" for k, v in df["Platform"].value_counts().items()])

    body = f"""Hi,

Here are your daily job results for {today}.

Total jobs found: {total}

Breakdown by platform:
{platform_summary}

Keywords searched: {", ".join(KEYWORDS)}
Location: {LOCATION}

The full list with apply links is attached as an Excel file.

Good luck with your search!
"""

    msg            = MIMEMultipart()
    msg["From"]    = GMAIL_USER
    msg["To"]      = RECIPIENT_EMAIL
    msg["Subject"] = f"Daily Jobs - {today} ({total} listings)"
    msg.attach(MIMEText(body, "plain"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(excel_bytes.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename=jobs_{today}.xlsx")
    msg.attach(part)

    print(f"Sending email to: {RECIPIENT_EMAIL}")
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_USER, RECIPIENT_EMAIL, msg.as_string())
    print("Email sent successfully.")

# -----------------------------------------------
# MAIN
# -----------------------------------------------
if __name__ == "__main__":
    print(f"Starting daily job search - {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"Keywords: {KEYWORDS}")
    print(f"Location: {LOCATION}")
    print(f"Sending results to: {RECIPIENT_EMAIL}\n")

    df = fetch_all_jobs()

    if df.empty:
        print("No jobs found. Nothing to send.")
    else:
        print(f"\nTotal jobs found: {len(df)}")
        print(df["Platform"].value_counts().to_string())
        excel = build_excel(df)
        send_email(df, excel)
        print("Done!")
