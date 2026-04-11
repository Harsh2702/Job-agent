import os
import requests
import smtplib
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# --- Config ---
ADZUNA_APP_ID = os.environ["ADZUNA_APP_ID"]
ADZUNA_APP_KEY = os.environ["ADZUNA_APP_KEY"]
GMAIL_USER = os.environ["GMAIL_USER"]
GMAIL_APP_PASSWORD = os.environ["GMAIL_APP_PASSWORD"]
RECIPIENT_EMAIL = "harshjoshi.3077@gmail.com"

KEYWORDS = ["Data Scientist", "Machine Learning Engineer", "AI Engineer"]
LOCATION = "germany"


# --- Fetch Jobs ---
def search_jobs(keyword):
    url = f"https://api.adzuna.com/v1/api/jobs/de/search/1"
    params = {
        "app_id": ADZUNA_APP_ID,
        "app_key": ADZUNA_APP_KEY,
        "what": keyword,
        "results_per_page": 20,
        "max_days_old": 1,
        "content-type": "application/json",
    }
    response = requests.get(url, params=params, timeout=30)
    print(f"Status code for '{keyword}': {response.status_code}")
    print(f"Raw response snippet: {response.text[:300]}")
    response.raise_for_status()
    return response.json().get("results", [])


def fetch_all_jobs():
    seen_ids = set()
    all_jobs = []

    for keyword in KEYWORDS:
        print(f"Searching: {keyword}")
        try:
            jobs = search_jobs(keyword)
        except Exception as e:
            print(f"Error fetching '{keyword}': {e}")
            continue

        for job in jobs:
            job_id = job.get("id")
            if job_id in seen_ids:
                continue
            seen_ids.add(job_id)

            location_str = job.get("location", {}).get("display_name", "")
            company = job.get("company", {}).get("display_name", "")
            date_raw = job.get("created", "")
            date_str = date_raw[:10] if date_raw else ""

            all_jobs.append({
                "Job Title": job.get("title", ""),
                "Company": company,
                "Location": location_str,
                "Date Posted": date_str,
                "Source": "Adzuna / Germany",
                "Apply Link": job.get("redirect_url", ""),
                "Keyword Match": keyword,
            })

    return all_jobs


# --- Create Excel ---
def create_excel(jobs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = ["Job Title", "Company", "Location", "Date Posted", "Source", "Apply Link", "Keyword Match"]

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="D9E1F2")

    for row_idx, job in enumerate(jobs, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=job.get(key, ""))
            cell.font = row_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    for row_idx, job in enumerate(jobs, 2):
        link = job.get("Apply Link", "")
        cell = ws.cell(row=row_idx, column=6)
        if link:
            cell.hyperlink = link
            cell.value = "Apply Here"
            cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    col_widths = [40, 30, 25, 15, 20, 15, 25]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"/tmp/jobs_{date_str}.xlsx"
    wb.save(filename)
    return filename


# --- Send Email ---
def send_email(filename, job_count):
    date_str = datetime.now().strftime("%d %b %Y")

    msg = MIMEMultipart()
    msg["From"] = GMAIL_USER
    msg["To"] = RECIPIENT_EMAIL
    msg["Subject"] = f"Daily Job Search - {date_str} ({job_count} jobs found)"

    body = f"""Hi Harsh,

Here are today's job listings for Germany.

Total unique jobs found: {job_count}
Keywords searched: {", ".join(KEYWORDS)}
Date: {date_str}

Open the attached Excel file to see all listings with apply links.

Good luck!
"""
    msg.attach(MIMEText(body, "plain"))

    with open(filename, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename=jobs_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
        msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.send_message(msg)

    print("Email sent successfully.")


def send_no_jobs_email():
    msg = MIMEMultipart()
    msg["From"] = GMAIL_USER
    msg["To"] = RECIPIENT_EMAIL
    msg["Subject"] = f"Daily Job Search - {datetime.now().strftime('%d %b %Y')} (no new jobs today)"

    body = "Hi Harsh,\n\nNo new jobs were found today for your keywords in Germany.\n\nThe search will run again tomorrow.\n"
    msg.attach(MIMEText(body, "plain"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.send_message(msg)

    print("No jobs found. Notification email sent.")


# --- Main ---
if __name__ == "__main__":
    print(f"Starting job search at {datetime.now()}")
    jobs = fetch_all_jobs()
    print(f"Total unique jobs found: {len(jobs)}")

    if jobs:
        filename = create_excel(jobs)
        send_email(filename, len(jobs))
    else:
        send_no_jobs_email()
